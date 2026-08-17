"""生成去雾+覆冰检测处理结果Excel报告"""
import os
import re
import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

INPUT_DIR = r"D:\dehaze_fusion\my_test\input"
OUTPUT_DIR = r"D:\dehaze_fusion\my_test\output"
COMPARE_DIR = r"D:\dehaze_fusion\my_test\compare"
EXCEL_PATH = r"D:\dehaze_fusion\my_test\processing_report.xlsx"


def compute_metrics(orig, dehazed):
    """计算去雾前后图像质量指标"""
    orig_gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)
    dehazed_gray = cv2.cvtColor(dehazed, cv2.COLOR_BGR2GRAY)

    orig_hsv = cv2.cvtColor(orig, cv2.COLOR_BGR2HSV)
    dehazed_hsv = cv2.cvtColor(dehazed, cv2.COLOR_BGR2HSV)

    # 亮度 (均值)
    orig_bright = np.mean(orig_gray)
    dehazed_bright = np.mean(dehazed_gray)

    # 对比度 (标准差)
    orig_contrast = np.std(orig_gray)
    dehazed_contrast = np.std(dehazed_gray)

    # 饱和度均值
    orig_sat = np.mean(orig_hsv[:, :, 1])
    dehazed_sat = np.mean(dehazed_hsv[:, :, 1])

    # 边缘密度 (Canny)
    orig_edges = cv2.Canny(orig_gray, 50, 150)
    dehazed_edges = cv2.Canny(dehazed_gray, 50, 150)
    orig_edge_density = np.sum(orig_edges > 0) / orig_gray.size * 100
    dehazed_edge_density = np.sum(dehazed_edges > 0) / dehazed_gray.size * 100

    # 暗通道 (去雾效果指标, 越低越清晰)
    def dark_channel(img, size=7):
        min_r = cv2.erode(img[:, :, 0], np.ones((size, size)))
        min_g = cv2.erode(img[:, :, 1], np.ones((size, size)))
        min_b = cv2.erode(img[:, :, 2], np.ones((size, size)))
        return np.minimum(np.minimum(min_r, min_g), min_b)
    orig_dc = np.mean(dark_channel(orig))
    dehazed_dc = np.mean(dark_channel(dehazed))

    # 信息熵
    def entropy(gray):
        hist = cv2.calcHist([gray], [0], None, [256], [0, 256])
        hist = hist.flatten() / hist.sum()
        hist = hist[hist > 0]
        return -np.sum(hist * np.log2(hist))
    orig_entropy = entropy(orig_gray)
    dehazed_entropy = entropy(dehazed_gray)

    return {
        '亮度_原图': round(orig_bright, 2),
        '亮度_去雾后': round(dehazed_bright, 2),
        '亮度提升': round(dehazed_bright - orig_bright, 2),
        '对比度_原图': round(orig_contrast, 2),
        '对比度_去雾后': round(dehazed_contrast, 2),
        '对比度提升': round(dehazed_contrast - orig_contrast, 2),
        '饱和度_原图': round(orig_sat, 2),
        '饱和度_去雾后': round(dehazed_sat, 2),
        '饱和度变化': round(dehazed_sat - orig_sat, 2),
        '边缘密度_原图': round(orig_edge_density, 2),
        '边缘密度_去雾后': round(dehazed_edge_density, 2),
        '暗通道_原图': round(orig_dc, 2),
        '暗通道_去雾后': round(dehazed_dc, 2),
        '信息熵_原图': round(orig_entropy, 4),
        '信息熵_去雾后': round(dehazed_entropy, 4),
    }


def analyze_ice_mask(mask_path):
    """分析覆冰掩码"""
    mask = cv2.imread(mask_path, cv2.IMREAD_GRAYSCALE)
    if mask is None:
        return None
    total = mask.size
    ice_pixels = np.sum(mask > 127)
    ice_ratio = ice_pixels / total * 100

    # 连通区域分析
    binary = (mask > 127).astype(np.uint8)
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(binary)
    ice_regions = max(0, num_labels - 1)  # 减去背景

    # 最大覆冰区域面积
    max_area = 0
    if ice_regions > 0:
        areas = stats[1:, cv2.CC_STAT_AREA]
        max_area = int(np.max(areas))

    return {
        '覆冰面积比(%)': round(ice_ratio, 2),
        '覆冰区域数': ice_regions,
        '最大覆冰区域(像素)': max_area,
    }


def parse_det_from_name(fname, output_dir):
    """从文件名推断检测结果文件路径 (支持YOLO和Mask R-CNN)"""
    base, ext = os.path.splitext(fname)
    maskrcnn_path = os.path.join(output_dir, f"{base}_maskrcnn{ext}")
    yolo_path = os.path.join(output_dir, f"{base}_yolo{ext}")
    if os.path.exists(maskrcnn_path):
        det_path, det_type = maskrcnn_path, 'Mask R-CNN'
    elif os.path.exists(yolo_path):
        det_path, det_type = yolo_path, 'YOLO'
    else:
        det_path, det_type = yolo_path, 'None'
    ice_path = os.path.join(output_dir, f"{base}_ice_mask{ext}")
    return det_path, det_type, ice_path


def get_image_files(directory, exts=('.jpg', '.jpeg', '.png', '.bmp', '.JPG')):
    files = []
    for f in os.listdir(directory):
        if f.endswith(exts):
            files.append(f)
    return sorted(files)


def main():
    input_files = get_image_files(INPUT_DIR)
    if not input_files:
        print("未找到输入图片")
        return

    print(f"分析 {len(input_files)} 张图片...")

    results = []
    for fname in input_files:
        orig_path = os.path.join(INPUT_DIR, fname)
        out_path = os.path.join(OUTPUT_DIR, fname)
        det_path, det_type, mask_path = parse_det_from_name(fname, OUTPUT_DIR)
        compare_path = os.path.join(COMPARE_DIR, f"compare_{fname}")

        orig = cv2.imread(orig_path)
        dehazed = cv2.imread(out_path)

        if orig is None or dehazed is None:
            print(f"  [跳过] {fname}")
            continue

        h, w = orig.shape[:2]
        metrics = compute_metrics(orig, dehazed)

        has_det = det_type != 'None'
        has_mask = os.path.exists(mask_path)
        has_compare = os.path.exists(compare_path)

        ice_data = analyze_ice_mask(mask_path) if has_mask else None

        row = {
            '文件名': fname,
            '分辨率': f"{w}x{h}",
            '文件大小(KB)': round(os.path.getsize(orig_path) / 1024, 1),
        }
        row.update(metrics)
        row['检测器'] = det_type if has_det else '无'
        row['目标检测'] = '有' if has_det else '无'
        row['对比图'] = '已生成' if has_compare else '未生成'

        if ice_data:
            row.update(ice_data)
            row['覆冰判定'] = '有覆冰'
        else:
            row['覆冰面积比(%)'] = 0
            row['覆冰区域数'] = 0
            row['最大覆冰区域(像素)'] = 0
            row['覆冰判定'] = '无覆冰'

        results.append(row)
        status = f"覆冰{ice_data['覆冰面积比(%)']:.1f}%" if ice_data else "无覆冰"
        print(f"  [完成] {fname} | {w}x{h} | {status}")

    # ==================== 生成Excel ====================
    wb = Workbook()

    # 样式定义
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cell_font = Font(name='微软雅黑', size=10)
    ice_fill = PatternFill(start_color='FFD6D6', end_color='FFD6D6', fill_type='solid')
    no_ice_fill = PatternFill(start_color='D6FFD6', end_color='D6FFD6', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # === Sheet1: 处理总览 ===
    ws1 = wb.active
    ws1.title = "处理总览"

    headers1 = ['文件名', '分辨率', '文件大小(KB)', '亮度提升', '对比度提升',
                '饱和度变化', '暗通道_原图', '暗通道_去雾后', '信息熵_去雾后',
                '检测器', '目标检测', '覆冰面积比(%)', '覆冰区域数', '覆冰判定', '对比图']

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, row in enumerate(results, 2):
        vals = [row.get(h, '') for h in headers1]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=col, value=v)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border
        if row.get('覆冰判定') == '有覆冰':
            for col in range(1, len(headers1) + 1):
                ws1.cell(row=i, column=col).fill = ice_fill
        else:
            for col in range(1, len(headers1) + 1):
                ws1.cell(row=i, column=col).fill = no_ice_fill

    col_widths1 = [42, 12, 14, 12, 12, 12, 14, 16, 14, 12, 10, 16, 12, 12, 12]
    for col, width in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = width
    ws1.row_dimensions[1].height = 30
    ws1.freeze_panes = 'A2'

    # === Sheet2: 去雾质量指标 ===
    ws2 = wb.create_sheet("去雾质量指标")
    headers2 = ['文件名', '分辨率',
                '亮度_原图', '亮度_去雾后', '亮度提升',
                '对比度_原图', '对比度_去雾后', '对比度提升',
                '饱和度_原图', '饱和度_去雾后', '饱和度变化',
                '边缘密度_原图', '边缘密度_去雾后',
                '暗通道_原图', '暗通道_去雾后',
                '信息熵_原图', '信息熵_去雾后']

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, row in enumerate(results, 2):
        vals = [row.get(h, '') for h in headers2]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=v)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border

    col_widths2 = [42, 12, 12, 14, 12, 14, 16, 14, 14, 16, 14, 16, 18, 14, 16, 14, 16]
    for col, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = 'A2'

    # === Sheet3: 覆冰检测详情 ===
    ws3 = wb.create_sheet("覆冰检测详情")
    headers3 = ['文件名', '分辨率', '检测器', '目标检测', '覆冰面积比(%)', '覆冰区域数',
                '最大覆冰区域(像素)', '覆冰判定', '暗通道_原图', '暗通道_去雾后',
                '边缘密度_去雾后']

    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, row in enumerate(results, 2):
        vals = [row.get(h, '') for h in headers3]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=col, value=v)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border
        if row.get('覆冰判定') == '有覆冰':
            for col in range(1, len(headers3) + 1):
                ws3.cell(row=i, column=col).fill = ice_fill
        else:
            for col in range(1, len(headers3) + 1):
                ws3.cell(row=i, column=col).fill = no_ice_fill

    col_widths3 = [42, 12, 12, 10, 16, 12, 18, 12, 14, 16, 18]
    for col, width in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(col)].width = width
    ws3.row_dimensions[1].height = 30
    ws3.freeze_panes = 'A2'

    # === Sheet4: 统计汇总 ===
    ws4 = wb.create_sheet("统计汇总")
    total = len(results)
    ice_count = sum(1 for r in results if r['覆冰判定'] == '有覆冰')
    no_ice_count = total - ice_count
    det_count = sum(1 for r in results if r.get('目标检测') == '有')
    det_types = [r['检测器'] for r in results if r.get('检测器', '无') != '无']
    det_type_summary = ', '.join(f"{t}: {det_types.count(t)}张" for t in sorted(set(det_types))) if det_types else '无'
    avg_bright_up = np.mean([r['亮度提升'] for r in results])
    avg_contrast_up = np.mean([r['对比度提升'] for r in results])
    avg_dc_orig = np.mean([r['暗通道_原图'] for r in results])
    avg_dc_dehazed = np.mean([r['暗通道_去雾后'] for r in results])
    avg_entropy = np.mean([r['信息熵_去雾后'] for r in results])
    ice_imgs = [r for r in results if r['覆冰判定'] == '有覆冰']
    avg_ice_ratio = np.mean([r['覆冰面积比(%)'] for r in ice_imgs]) if ice_imgs else 0

    stats = [
        ['统计项目', '数值'],
        ['总图片数', total],
        ['有覆冰图片数', ice_count],
        ['无覆冰图片数', no_ice_count],
        ['有检测图片数', det_count],
        ['检测器分布', det_type_summary],
        ['平均亮度提升', round(avg_bright_up, 2)],
        ['平均对比度提升', round(avg_contrast_up, 2)],
        ['平均暗通道(原图)', round(avg_dc_orig, 2)],
        ['平均暗通道(去雾后)', round(avg_dc_dehazed, 2)],
        ['平均信息熵(去雾后)', round(avg_entropy, 4)],
        ['平均覆冰面积比(有覆冰图片)', round(avg_ice_ratio, 2)],
        ['使用模型', 'M4 (ITL覆冰感知损失)'],
        ['去雾算法', 'IceWave-DehazeFormer-S + HA-WFE v2 + CLIP蒸馏'],
        ['目标检测', det_type_summary if det_types else 'YOLOv8自定义 + HSV颜色+纹理+轮廓形状过滤'],
        ['覆冰检测', 'HSV颜色+纹理+轮廓形状过滤 + 检测走廊约束'],
    ]

    for row_idx, (label, val) in enumerate(stats, 1):
        for col_idx, val2 in enumerate([label, val], 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val2)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = cell_font

    ws4.column_dimensions['A'].width = 32
    ws4.column_dimensions['B'].width = 50
    ws4.row_dimensions[1].height = 30

    # 保存
    wb.save(EXCEL_PATH)
    print(f"\nExcel报告已保存: {EXCEL_PATH}")
    print(f"包含 {len(results)} 张图片的数据, 4个工作表: 处理总览/去雾质量指标/覆冰检测详情/统计汇总")


if __name__ == "__main__":
    main()
